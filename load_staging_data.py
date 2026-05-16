from pathlib import Path

import pyodbc
from openpyxl import load_workbook

from utils import get_sql_config, create_connection_string


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = PROJECT_ROOT / "data" / "raw" / "raw_data_source.xlsx"
CONFIG_PATH = PROJECT_ROOT / "sql_server_config.cfg"


SHEET_TO_TABLE = {
    "Categories": "Staging_Categories",
    "Customers": "Staging_Customers",
    "Employees": "Staging_Employees",
    "OrderDetails": "Staging_Order_Details",
    "Orders": "Staging_Orders",
    "Products": "Staging_Products",
    "Region": "Staging_Region",
    "Shippers": "Staging_Shippers",
    "Suppliers": "Staging_Suppliers",
    "Territories": "Staging_Territories",
}


def load_excel_to_staging(
    excel_path: Path = DEFAULT_EXCEL_PATH,
    config_path: Path = CONFIG_PATH,
):
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Raw Excel file not found: {excel_path}. "
            "Place raw_data_source.xlsx in data/raw/."
        )

    sql_config = get_sql_config(str(config_path))
    connection_string = create_connection_string(sql_config)

    workbook = load_workbook(excel_path, data_only=True)
    connection = pyodbc.connect(connection_string)
    cursor = connection.cursor()

    try:
        for sheet_name, table_name in SHEET_TO_TABLE.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Missing sheet in Excel file: {sheet_name}")

            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                print(f"Skipping empty sheet: {sheet_name}")
                continue

            columns = list(rows[0])
            data_rows = rows[1:]

            placeholders = ", ".join(["?"] * len(columns))
            column_sql = ", ".join(f"[{column}]" for column in columns)

            cursor.execute(f"DELETE FROM dbo.{table_name};")

            insert_sql = (
                f"INSERT INTO dbo.{table_name} ({column_sql}) "
                f"VALUES ({placeholders})"
            )

            loaded_rows = 0
            for row in data_rows:
                cursor.execute(insert_sql, row)
                loaded_rows += 1

            print(f"Loaded {loaded_rows} rows into {table_name}")

        connection.commit()
        return {"success": True}

    except Exception:
        connection.rollback()
        raise

    finally:
        connection.close()


if __name__ == "__main__":
    result = load_excel_to_staging()
    print(result)
